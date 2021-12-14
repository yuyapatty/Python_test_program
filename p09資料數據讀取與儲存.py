# -*- coding: utf-8 -*-
"""
Created on Fri Jul  2 15:11:28 2021

@author: YY
"""
# 資料數據讀取與儲存
import openpyxl
import json
import csv
f1 = open('ex1.txt', 'r', encoding='utf-8')
f2 = open('ex1-1.txt', 'w', encoding='utf-8')
# print(f1.read(5))
content = f1.readlines()
i = 1
for item in content:
    f2.write('%02d' % i)
    f2.write(item)
    i += 1

f1.close()
f1.close()

# 讀取csv
f1 = open('ex2.csv', 'r')
rows = csv.reader(f1)
for row in rows:
    print(row)

# DictReader
f1 = open('ex2.csv', 'r')
rows = csv.DictReader(f1)
for row in rows:
    print(row['姓名'], row['學號'])

# 讀取csv
with open('ex2.csv', 'r', newline='') as f1:
    data = csv.reader(f1)
    for row in data:
        print(row)

# 讀取新增寫入
with open('ex2.csv', 'r', newline='') as f1:
    with open('ex2_1.csv', 'w', newline='') as f2:
        data = csv.reader(f1)
        writer = csv.writer(f2)
        for row in data:
            if row[0] == '學號':
                row.append('總分')
            else:
                row.append(int(row[2])+int(row[3])+int(row[4]))
            writer.writerow(row)

# 使用字典方法，讀取新增寫入
with open('ex2.csv', 'r', newline='') as f1:
    with open('ex2_2.csv', 'w', newline='') as f2:
        data = csv.DictReader(f1)
        fields = ['學號', '姓名', '國文', '英文', '數學', '總分']
        writer = csv.DictWriter(f2, fieldnames=fields)
        writer.writeheader()
        for row in data:
            row['總分'] = int(row['國文'])+int(row['英文'])+int(row['數學'])
            writer.writerow(row)

# 下載檔案讀取

with open('aqi.csv', 'r', newline='', encoding='utf-8-sig') as f1:
    # BOM，前面會因為是utf-8的格式而要加上sig
    data = csv.DictReader(f1)
    for row in data:
        if row['SiteName'] == '橋頭':
            print(row['SiteName'], row['AQI'], row['Status'])

# jason

with open('tcbnb.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    for item in data:
        print(item['中文名稱'], item['地址'])

# jason轉csv

with open('tcbnb.json', 'r', encoding='utf-8') as f1:
    with open('tcbnb.csv', 'w', encoding='utf-8') as f2:
        data = json.load(f1)
        writer = csv.writer(f2)
        writer.writerow(['中文名稱', '地址', '電話或手機'])
        for item in data:
            new = list()
            new.append(item['中文名稱'])
            new.append(item['地址'])
            new.append(item['電話或手機'])
            writer.writerow(new)

# Excel

workbook = openpyxl.Workbook()
sheet = workbook.worksheets[0]
sheet['A1'] = '一年甲班'
listtitle = ['座號', '姓名', '國文', '英文', '數學']
sheet.append(listtitle)
listdatas = [[1, '葉大雄', 65, 62, 40],
             [2, '陳靜香', 85, 90, 87],
             [3, '王聰明', 92, 90, 95]]
for listdata in listdatas:
    sheet.append(listdata)
workbook.save('test.xlsx')

# Excel

wb = openpyxl.Workbook()
sheet = wb.worksheets[0]
sheet['A1'] = '新尖兵訓練班'
sheet.append(['學號', '姓名', '電話'])
sheet.append(['1', 'David', '0912345678'])
sheet.append(['2', 'Mary', '015555555'])
sheet.append(['3', 'Tony', '0935888000'])
wb.save('myexcel.xlsx')

# Excel

wb = openpyxl.load_workbook('myexcel.xlsx')
sheet = wb.worksheets[0]
# sheet=wb.worksheets[0]
sheet['C4'] = '0988123888'
sheet.append(['4', 'Amy', '0925653267'])
wb.save('myexcel.xlsx')

# Excel，印出來

wb = openpyxl.load_workbook('myexcel.xlsx')
sheet = wb['Sheet']
print(sheet['A1'].value)
print(sheet.cell(row=1, column=1).value)
print(sheet.max_row, sheet.max_column)
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value != None:
            print("%-6s" % sheet.cell(row=i, column=j).value, end=" ")
        print()
